#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import os
import docx
from openpyxl import Workbook
from openpyxl.styles import Border, Side


class StuInfo:
    def __init__(self):
        self.id = ""  # 学号
        self.name = ""  # 姓名
        self.major = ""  # 专业
        self.advisor = ""  # 指导教师
        self.subject = ""  # 毕设题目
        self.scores = {"成绩": [], "评阅评语": [], "指导教师评价": []}  # 评分


def read_students(folder_path: str = "all_docs"):
    # 定义空列表，用于存放符合要求的内容
    stu_list = []
    for folder in os.listdir(folder_path):
        if folder.startswith("A2204") and os.path.isdir(
            os.path.join(folder_path, folder)
        ):
            stu = StuInfo()
            for document in os.listdir(os.path.join(folder_path, folder)):
                if document.endswith(".docx") and not document.startswith("~$"):
                    if "成绩表" in document:
                        mode = "成绩"
                    elif "评阅评语表" in document:
                        mode = "评阅评语"
                    elif "指导教师评价表" in document:
                        mode = "指导教师评价"
                    else:
                        break
                    doc_path = os.path.join(folder_path, folder, document)
                    doc = docx.Document(doc_path)
                    # 第二步：整理并归纳成方便读取的形式
                    for table in doc.tables:
                        done = False
                        for i, row in enumerate(table.rows):
                            for j, cell in enumerate(row.cells):
                                if (
                                    cell.text.strip() == "题目"
                                    and table.cell(i, j + 1).text.strip() != "题目"
                                ):
                                    stu.subject = table.cell(i, j + 1).text.strip()
                                elif (
                                    cell.text.strip() == "学生姓名"
                                    and table.cell(i, j + 1).text.strip() != "学生姓名"
                                ):
                                    stu.name = table.cell(i, j + 1).text.strip()
                                elif (
                                    cell.text.strip() == "班级学号"
                                    and table.cell(i, j + 1).text.strip() != "班级学号"
                                ):
                                    stu.id = table.cell(i, j + 1).text.strip()
                                elif (
                                    cell.text.strip() == "专业"
                                    and table.cell(i, j + 1).text.strip() != "专业"
                                ):
                                    stu.major = table.cell(i, j + 1).text.strip()
                                elif (
                                    cell.text.strip() == "指导教师"
                                    and table.cell(i, j + 1).text.strip() != "指导教师"
                                ):
                                    stu.advisor = table.cell(i, j + 1).text.strip()
                                elif (
                                    cell.text.strip() == "评分"
                                    and table.cell(i, j - 1).text.strip() == "总分"
                                ):
                                    if mode == "成绩":
                                        end = i + 5
                                    elif mode == "评阅评语":
                                        end = i + 7
                                    elif mode == "指导教师评价":
                                        end = i + 8
                                    else:
                                        break
                                    for k in range(i + 1, end):
                                        stu.scores[mode].append(
                                            float(table.cell(k, j).text.strip())
                                        )
                                    done = True
                                    break
                            if done:
                                break
            stu_list.append(stu)
    return stu_list


# 第三步：将整理好的数据写入excel表格
def write_excel(stu_list: list[StuInfo], file_path: str = "all_docs/output.xlsx"):
    wb = Workbook()
    ws = wb.active
    # 表头
    ws.append(["姓名", "学号", "专业", "指导教师", "毕设题目", "成绩"])
    ws.append(
        [
            "",
            "",
            "",
            "",
            "",
            "指导教师评价",
            "",
            "",
            "",
            "",
            "",
            "",
            "评阅评语",
            "",
            "",
            "",
            "",
            "",
            "成绩",
            "",
            "",
            "",
            "",
        ]
    )
    ws.append(
        [
            "",
            "",
            "",
            "",
            "",
            "开题",
            "外文",
            "设计",
            "创新",
            "撰写",
            "态度",
            "综合",
            "开题",
            "外文",
            "设计",
            "创新",
            "撰写",
            "综合",
            "设计",
            "创新",
            "答辩",
            "综合",
        ]
    )
    # 合并单元格
    ws.merge_cells("A1:A3")
    ws.merge_cells("B1:B3")
    ws.merge_cells("C1:C3")
    ws.merge_cells("D1:D3")
    ws.merge_cells("E1:E3")
    ws.merge_cells("F1:V1")
    ws.merge_cells("F2:L2")
    ws.merge_cells("M2:R2")
    ws.merge_cells("S2:V2")
    # 表格内容
    for stu in stu_list:
        data = [stu.name, stu.id, stu.major, stu.advisor, stu.subject]
        for mode in ["指导教师评价", "评阅评语", "成绩"]:
            data.extend(stu.scores[mode])
        ws.append(data)
    # 设置边框
    for row in ws.rows:
        for cell in row:
            if cell.row == 1:
                if cell.column == 1:
                    cell.border = Border(
                        left=Side(border_style="thin", color="000000"),
                        top=Side(border_style="thin", color="000000"),
                    )
                elif cell.column == ws.max_column:
                    cell.border = Border(
                        right=Side(border_style="thin", color="000000"),
                        top=Side(border_style="thin", color="000000"),
                    )
                else:
                    cell.border = Border(
                        top=Side(border_style="thin", color="000000"),
                    )
            elif cell.row == ws.max_row:
                if cell.column == 1:
                    cell.border = Border(
                        left=Side(border_style="thin", color="000000"),
                        bottom=Side(border_style="thin", color="000000"),
                    )
                elif cell.column == ws.max_column:
                    cell.border = Border(
                        right=Side(border_style="thin", color="000000"),
                        bottom=Side(border_style="thin", color="000000"),
                    )
                else:
                    cell.border = Border(
                        bottom=Side(border_style="thin", color="000000"),
                    )
            else:
                if cell.column == 1:
                    cell.border = Border(
                        left=Side(border_style="thin", color="000000"),
                    )
                elif cell.column == ws.max_column:
                    cell.border = Border(
                        right=Side(border_style="thin", color="000000"),
                    )

    # 保存文件
    wb.save(file_path)
    wb.close()


def run():
    write_excel(read_students("all_docs"), "all_docs/output.xlsx")


if __name__ == "__main__":
    run()
